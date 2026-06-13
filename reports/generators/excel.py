"""
Parliament IT Inventory — Excel report generators (openpyxl).

Each public function returns bytes suitable for an HttpResponse.
All reports include Parliament Blue headers and alternating row shading.
"""

import io
from datetime import date, timedelta

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.columns import (
    ASSET_HISTORY_COLS,
    HOLDER_ASSIGNMENTS_COLS,
    INVENTORY_COLS,
    LIFECYCLE_COLS,
    TRANSFER_LOG_COLS,
    WARRANTY_COLS,
)

# ── Brand colours ─────────────────────────────────────────────────────────────
_FILL_HDR   = PatternFill("solid", fgColor="0076A7")
_FILL_ALT   = PatternFill("solid", fgColor="E8F3F8")
_FONT_HDR   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_FONT_TITLE = Font(name="Calibri", bold=True, color="0076A7", size=13)
_FONT_ORG   = Font(name="Calibri", italic=True, color="777777", size=9)
_ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=False)
_ALIGN_L    = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# ── Per-report column → default width maps ────────────────────────────────────
_INV_WIDTHS: dict[str, int] = {
    "asset_tag": 14, "category": 14, "type": 16, "brand": 14, "model": 22,
    "serial_no": 16, "cpu": 24, "ram": 14, "storage": 16, "display": 10,
    "status": 12, "storage_location": 24, "current_holder": 28,
    "holder_type": 12, "assigned_since": 14, "purchase_date": 14,
    "warranty_expiry": 16, "amc_expiry": 14,
}
_TLOG_WIDTHS: dict[str, int] = {
    "transfer_date": 18, "asset_tag": 14, "category": 14, "type": 16, "brand": 14,
    "model": 22, "assigned_to": 28, "holder_type": 12, "designation": 32,
    "status": 14, "performed_by": 22, "batch_ref": 14, "notes": 32,
}
_LC_WIDTHS: dict[str, int] = {
    "date": 18, "asset_tag": 14, "category": 14, "type": 16, "brand": 14,
    "model": 22, "event": 22, "old_status": 14, "new_status": 14,
    "notes": 36, "performed_by": 22,
}
_WTY_WIDTHS: dict[str, int] = {
    "asset_tag": 14, "category": 14, "type": 16, "brand": 14, "model": 22,
    "status": 12, "current_holder": 28, "warranty_expiry": 16, "warranty_days": 10,
    "amc_expiry": 16, "amc_days": 10,
}
_HOLD_WIDTHS: dict[str, int] = {
    "holder": 28, "holder_type": 10, "designation": 34, "department": 28,
    "asset_tag": 14, "category": 14, "asset_type": 16, "brand": 14,
    "model": 22, "status": 12, "assigned_since": 16,
}
_HIST_WIDTHS: dict[str, int] = {
    "assigned_to": 28, "holder_type": 10, "designation": 34, "department": 26,
    "from_date": 16, "to_date": 16, "days": 8, "performed_by": 22,
    "batch_ref": 12, "notes": 34,
}


# ── Private helpers ───────────────────────────────────────────────────────────

def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _title_block(ws, title: str, subtitle: str = "") -> int:
    """Write organisation + report title; return the row index for the header row."""
    ws["A1"] = "Bangladesh Parliament Secretariat · IT Inventory"
    ws["A1"].font = _FONT_ORG
    ws["A2"] = title
    ws["A2"].font = _FONT_TITLE
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 22
    if subtitle:
        ws["A3"] = subtitle
        ws["A3"].font = Font(name="Calibri", italic=True, color="777777", size=9)
        ws.row_dimensions[3].height = 14
        return 5
    return 4


def _header_row(ws, row: int, cols: list[str]) -> None:
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill = _FILL_HDR
        cell.font = _FONT_HDR
        cell.alignment = _ALIGN_CTR
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _data_row(ws, row: int, values: list, alt: bool = False) -> None:
    fill = _FILL_ALT if alt else None
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.alignment = _ALIGN_L
        if fill:
            cell.fill = fill


def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _date_s(d) -> str:
    if d is None:
        return ""
    return d.strftime("%d %b %Y") if hasattr(d, "strftime") else str(d)


def _dt_s(dt) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d %b %Y %H:%M") if hasattr(dt, "strftime") else str(dt)


def _eff_cols(columns: list[str] | None, col_list: list[tuple[str, str]]) -> list[str]:
    """Return effective column keys: caller-supplied subset or all keys."""
    return columns if columns else [k for k, _ in col_list]


def _labels_for(keys: list[str], col_list: list[tuple[str, str]]) -> list[str]:
    label_map = {k: lbl for k, lbl in col_list}
    return [label_map[k] for k in keys]


def _widths_for(keys: list[str], width_map: dict[str, int]) -> list[int]:
    return [width_map.get(k, 14) for k in keys]


# ── Public report functions ───────────────────────────────────────────────────

def inventory_excel(
    filters: dict | None = None,
    columns: list[str] | None = None,
) -> bytes:
    """Current inventory: all non-deleted assets with live holder information."""
    from assets.models import AssetItem
    from assignments.models import Assignment

    filters = filters or {}
    qs = AssetItem.objects.filter(is_deleted=False).select_related(
        "asset_type__category", "storage_location",
    ).order_by("asset_tag")

    if filters.get("status"):
        qs = qs.filter(status=filters["status"])
    if filters.get("category"):
        qs = qs.filter(asset_type__category_id=filters["category"])
    if filters.get("type"):
        qs = qs.filter(asset_type_id=filters["type"])
    if filters.get("ram_type"):
        qs = qs.filter(specifications__ram__type=filters["ram_type"])
    if filters.get("os_name"):
        qs = qs.filter(specifications__os__name=filters["os_name"])
    if filters.get("os_licensed"):
        qs = qs.filter(specifications__os__licensed=filters["os_licensed"])

    # Advanced spec filters (CPU/RAM/storage/display) — applied in Python.
    from reports import spec_filters
    spec_f = spec_filters.parse_spec_filters(filters)
    assets = spec_filters.filter_assets(qs, spec_f) if spec_filters.is_active(spec_f) else list(qs)

    pks = [a.pk for a in assets]
    active_map = {
        a.asset_id: a
        for a in Assignment.objects.filter(
            asset_id__in=pks, returned_at__isnull=True,
        ).select_related("assignee__employee", "assignee__mp", "assignee__office")
    }

    eff = _eff_cols(columns, INVENTORY_COLS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    hrow = _title_block(ws, "Current IT Asset Inventory", f"Generated: {_date_s(date.today())}")
    _header_row(ws, hrow, _labels_for(eff, INVENTORY_COLS))
    _col_widths(ws, _widths_for(eff, _INV_WIDTHS))

    for i, asset in enumerate(assets):
        asgn = active_map.get(asset.pk)
        if asgn:
            holder = asgn.assignee.display_name
            htype  = asgn.assignee.get_assignee_type_display()
            since  = _date_s(asgn.assigned_at.date())
        else:
            holder = htype = since = ""

        specs = asset.specifications or {}
        row_dict = {
            "asset_tag":        asset.asset_tag,
            "category":         asset.asset_type.category.name,
            "type":             asset.asset_type.name,
            "brand":            asset.brand,
            "model":            asset.model_name,
            "serial_no":        asset.serial_number,
            "cpu":              spec_filters.fmt_cpu(specs),
            "ram":              spec_filters.fmt_ram(specs),
            "storage":          spec_filters.fmt_storage(specs),
            "display":          spec_filters.fmt_display(specs),
            "status":           asset.get_status_display(),
            "storage_location": asset.storage_location.full_path if asset.storage_location else "",
            "current_holder":   holder,
            "holder_type":      htype,
            "assigned_since":   since,
            "purchase_date":    _date_s(asset.purchase_date),
            "warranty_expiry":  _date_s(asset.warranty_expiry),
            "amc_expiry":       _date_s(asset.amc_expiry),
        }
        _data_row(ws, hrow + 1 + i, [row_dict[k] for k in eff], alt=(i % 2 == 1))

    return _wb_bytes(wb)


def transfer_log_excel(
    date_from=None,
    date_to=None,
    columns: list[str] | None = None,
) -> bytes:
    """All assignment records, optionally filtered by date range."""
    from assignments.models import Assignment

    qs = Assignment.objects.select_related(
        "asset__asset_type__category",
        "assignee__employee", "assignee__mp", "assignee__office",
        "performed_by", "batch",
    ).order_by("-assigned_at")

    if date_from:
        qs = qs.filter(assigned_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(assigned_at__date__lte=date_to)

    qs = qs[:5000]

    subtitle = ""
    if date_from or date_to:
        d1 = _date_s(date_from) if date_from else "Start"
        d2 = _date_s(date_to)   if date_to   else "Today"
        subtitle = f"Period: {d1} – {d2}"

    eff = _eff_cols(columns, TRANSFER_LOG_COLS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transfer Log"
    hrow = _title_block(ws, "Asset Transfer Log", subtitle)
    _header_row(ws, hrow, _labels_for(eff, TRANSFER_LOG_COLS))
    _col_widths(ws, _widths_for(eff, _TLOG_WIDTHS))

    for i, asgn in enumerate(qs):
        asset = asgn.asset
        snap  = asgn.holder_snapshot or {}
        status = "Active" if asgn.returned_at is None else f"Returned {_date_s(asgn.returned_at.date())}"

        row_dict = {
            "transfer_date": _dt_s(asgn.assigned_at),
            "asset_tag":     asset.asset_tag,
            "category":      asset.asset_type.category.name,
            "type":          asset.asset_type.name,
            "brand":         asset.brand,
            "model":         asset.model_name,
            "assigned_to":   snap.get("display_name", ""),
            "holder_type":   snap.get("assignee_type", ""),
            "designation":   snap.get("designation", ""),
            "status":        status,
            "performed_by":  asgn.performed_by.get_full_name() or asgn.performed_by.username,
            "batch_ref":     asgn.batch.reference if asgn.batch_id else "",
            "notes":         asgn.notes,
        }
        _data_row(ws, hrow + 1 + i, [row_dict[k] for k in eff], alt=(i % 2 == 1))

    return _wb_bytes(wb)


def lifecycle_events_excel(
    date_from=None,
    date_to=None,
    event_type=None,
    columns: list[str] | None = None,
) -> bytes:
    """Lifecycle events log, optionally filtered by date and event type."""
    from assets.models import AssetItem
    from lifecycle.models import LifecycleEvent

    qs = LifecycleEvent.objects.select_related(
        "asset__asset_type__category", "performed_by",
    ).order_by("-occurred_at")

    if date_from:
        qs = qs.filter(occurred_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(occurred_at__date__lte=date_to)
    if event_type:
        qs = qs.filter(event_type=event_type)

    qs = qs[:5000]

    eff = _eff_cols(columns, LIFECYCLE_COLS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lifecycle Events"
    hrow = _title_block(ws, "Lifecycle Events Log")
    _header_row(ws, hrow, _labels_for(eff, LIFECYCLE_COLS))
    _col_widths(ws, _widths_for(eff, _LC_WIDTHS))

    for i, ev in enumerate(qs):
        try:
            old_label = AssetItem.Status(ev.old_status).label
        except ValueError:
            old_label = ev.old_status
        try:
            new_label = AssetItem.Status(ev.new_status).label
        except ValueError:
            new_label = ev.new_status

        row_dict = {
            "date":         _dt_s(ev.occurred_at),
            "asset_tag":    ev.asset.asset_tag,
            "category":     ev.asset.asset_type.category.name,
            "type":         ev.asset.asset_type.name,
            "brand":        ev.asset.brand,
            "model":        ev.asset.model_name,
            "event":        ev.get_event_type_display(),
            "old_status":   old_label,
            "new_status":   new_label,
            "notes":        ev.note,
            "performed_by": ev.performed_by.get_full_name() or ev.performed_by.username,
        }
        _data_row(ws, hrow + 1 + i, [row_dict[k] for k in eff], alt=(i % 2 == 1))

    return _wb_bytes(wb)


def warranty_expiry_excel(
    days: int = 90,
    columns: list[str] | None = None,
) -> bytes:
    """Assets with warranty or AMC expiring within `days` days."""
    from assets.models import AssetItem
    from assignments.models import Assignment
    from django.db.models import Q

    today   = date.today()
    horizon = today + timedelta(days=days)

    qs = AssetItem.objects.filter(
        is_deleted=False,
    ).filter(
        Q(warranty_expiry__lte=horizon) | Q(amc_expiry__lte=horizon)
    ).select_related("asset_type__category", "storage_location").order_by(
        "warranty_expiry", "amc_expiry",
    )

    pks = list(qs.values_list("pk", flat=True))
    active_map = {
        a.asset_id: a
        for a in Assignment.objects.filter(
            asset_id__in=pks, returned_at__isnull=True,
        ).select_related("assignee__employee", "assignee__mp", "assignee__office")
    }

    eff = _eff_cols(columns, WARRANTY_COLS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Warranty AMC Expiry"
    hrow = _title_block(
        ws, "Warranty / AMC Expiry Report",
        f"Assets expiring within {days} days — as of {_date_s(today)}",
    )
    _header_row(ws, hrow, _labels_for(eff, WARRANTY_COLS))
    _col_widths(ws, _widths_for(eff, _WTY_WIDTHS))

    for i, asset in enumerate(qs):
        asgn  = active_map.get(asset.pk)
        holder = asgn.assignee.display_name if asgn else ""
        wty_d  = (asset.warranty_expiry - today).days if asset.warranty_expiry else None
        amc_d  = (asset.amc_expiry      - today).days if asset.amc_expiry      else None

        row_dict = {
            "asset_tag":       asset.asset_tag,
            "category":        asset.asset_type.category.name,
            "type":            asset.asset_type.name,
            "brand":           asset.brand,
            "model":           asset.model_name,
            "status":          asset.get_status_display(),
            "current_holder":  holder,
            "warranty_expiry": _date_s(asset.warranty_expiry),
            "warranty_days":   "" if wty_d is None else wty_d,
            "amc_expiry":      _date_s(asset.amc_expiry),
            "amc_days":        "" if amc_d is None else amc_d,
        }
        _data_row(ws, hrow + 1 + i, [row_dict[k] for k in eff], alt=(i % 2 == 1))

    return _wb_bytes(wb)


def holder_assignments_excel(
    holder_type: str | None = None,
    columns: list[str] | None = None,
) -> bytes:
    """Current active assignments grouped by holder, optionally filtered by type."""
    from assignments.models import Assignment
    from assignees.models import AssigneeType

    qs = (
        Assignment.objects.filter(returned_at__isnull=True)
        .select_related(
            "asset__asset_type__category",
            "assignee__employee", "assignee__mp", "assignee__office",
        )
        .order_by(
            "assignee__assignee_type",
            "assignee__employee__name_en",
            "assignee__mp__name_en",
            "assignee__office__name_en",
            "asset__asset_tag",
        )
    )

    if holder_type in (t.value for t in AssigneeType):
        qs = qs.filter(assignee__assignee_type=holder_type)

    subtitle = f"Holder type: {holder_type}" if holder_type else "All holder types"
    eff = _eff_cols(columns, HOLDER_ASSIGNMENTS_COLS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Holder Assignments"
    hrow = _title_block(
        ws, "Current Holder Assignments",
        f"{subtitle}  ·  Generated: {_date_s(date.today())}",
    )
    _header_row(ws, hrow, _labels_for(eff, HOLDER_ASSIGNMENTS_COLS))
    _col_widths(ws, _widths_for(eff, _HOLD_WIDTHS))

    for i, asgn in enumerate(qs):
        snap = asgn.holder_snapshot or {}
        row_dict = {
            "holder":        asgn.assignee.display_name,
            "holder_type":   asgn.assignee.get_assignee_type_display(),
            "designation":   snap.get("designation", ""),
            "department":    snap.get("department", ""),
            "asset_tag":     asgn.asset.asset_tag,
            "category":      asgn.asset.asset_type.category.name,
            "asset_type":    asgn.asset.asset_type.name,
            "brand":         asgn.asset.brand,
            "model":         asgn.asset.model_name,
            "status":        asgn.asset.get_status_display(),
            "assigned_since": _date_s(asgn.assigned_at.date()),
        }
        _data_row(ws, hrow + 1 + i, [row_dict[k] for k in eff], alt=(i % 2 == 1))

    return _wb_bytes(wb)


def asset_history_excel(
    asset_pk: int,
    columns: list[str] | None = None,
) -> bytes:
    """Full assignment history for one asset."""
    from assets.models import AssetItem
    from assignments.models import Assignment

    asset = AssetItem.objects.select_related("asset_type__category").get(pk=asset_pk)
    history = Assignment.objects.filter(asset=asset).select_related(
        "performed_by", "batch",
    ).order_by("-assigned_at")

    eff = _eff_cols(columns, ASSET_HISTORY_COLS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asset History"
    hrow = _title_block(
        ws, f"Asset History — {asset.asset_tag}",
        f"{asset.brand} {asset.model_name}  ·  {asset.asset_type.name}",
    )

    # "#" SL column is always present; selected cols follow
    _header_row(ws, hrow, ["#"] + _labels_for(eff, ASSET_HISTORY_COLS))
    _col_widths(ws, [4] + _widths_for(eff, _HIST_WIDTHS))

    for i, asgn in enumerate(history, 1):
        snap = asgn.holder_snapshot or {}
        if asgn.returned_at:
            to_date = _date_s(asgn.returned_at.date())
            days    = (asgn.returned_at - asgn.assigned_at).days
        else:
            to_date = "Current"
            days    = (timezone.now() - asgn.assigned_at).days

        row_dict = {
            "assigned_to":  snap.get("display_name", ""),
            "holder_type":  snap.get("assignee_type", ""),
            "designation":  snap.get("designation", ""),
            "department":   snap.get("department", ""),
            "from_date":    _date_s(asgn.assigned_at.date()),
            "to_date":      to_date,
            "days":         days,
            "performed_by": asgn.performed_by.get_full_name() or asgn.performed_by.username,
            "batch_ref":    asgn.batch.reference if asgn.batch_id else "",
            "notes":        asgn.notes,
        }
        _data_row(ws, hrow + i, [i] + [row_dict[k] for k in eff], alt=(i % 2 == 0))

    return _wb_bytes(wb)
