"""
Excel-based bulk import for AssetItem.

Workflow: generate_template → user fills in → validate → preview → execute.
"""
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from assets.models import AssetItem, AssetType
from locations.models import Location

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FIXED_COLUMNS: list[str] = [
    "asset_tag",
    "serial_number",
    "brand",
    "model_name",
    "purchase_date",
    "purchase_order",
    "supplier",
    "purchase_cost",
    "warranty_expiry",
    "amc_expiry",
    "storage_location_path",
    "notes",
]

COLUMN_DESCRIPTIONS: dict[str, str] = {
    "asset_tag": "Optional. Leave blank for auto-generation (e.g. LAP-2026-0001).",
    "serial_number": "Hardware serial number. Strongly recommended.",
    "brand": "REQUIRED. Manufacturer name (e.g. Dell, HP, Cisco).",
    "model_name": "REQUIRED. Model number or name (e.g. Latitude 5540).",
    "purchase_date": "Format: YYYY-MM-DD (e.g. 2024-01-15). Optional.",
    "purchase_order": "Purchase order / work order number. Optional.",
    "supplier": "Vendor / supplier name. Optional.",
    "purchase_cost": "Numeric value only, no currency symbol (e.g. 45000.00). Optional.",
    "warranty_expiry": "Format: YYYY-MM-DD. Optional.",
    "amc_expiry": "Annual Maintenance Contract expiry. Format: YYYY-MM-DD. Optional.",
    "storage_location_path": (
        "Copy-paste exactly from the 'Valid Locations' sheet. "
        "Use the → separator (e.g. Parliament Bhaban → 3rd Floor → NOC Room). Optional."
    ),
    "notes": "Any additional notes. Optional.",
}

_HEADER_FILL = PatternFill(start_color="006633", end_color="006633", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_EXAMPLE_FONT = Font(italic=True, color="595959", name="Calibri", size=10)
_INSTR_HEADER_FONT = Font(bold=True, name="Calibri", size=12, color="006633")
_INSTR_FONT = Font(name="Calibri", size=10)

SESSION_KEY_ROWS = "excel_import_validated_rows"
SESSION_KEY_TYPE = "excel_import_asset_type_id"
SESSION_KEY_COLS = "excel_import_columns"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_type_prefix(asset_type_name: str) -> str:
    """First 3 uppercase alpha characters of the type name, e.g. 'LAPTOP' → 'LAP'."""
    cleaned = re.sub(r"[^A-Za-z]", "", asset_type_name).upper()
    return cleaned[:3].ljust(3, "X")


def _normalize_location_path(raw: str) -> str:
    """Normalise separators and whitespace for matching."""
    return re.sub(r"\s*[→>]\s*", "→", raw).strip().lower()


def _build_location_lookup() -> dict[str, int]:
    """Return {normalised_full_path: location_pk} for all active locations."""
    lookup: dict[str, int] = {}
    for loc in Location.objects.filter(is_active=True).select_related("parent__parent"):
        key = _normalize_location_path(loc.full_path)
        lookup[key] = loc.pk
    return lookup


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    # Try multiple formats; Excel serial dates become "YYYY-MM-DD HH:MM:SS" strings
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date '{value}'. Use YYYY-MM-DD (e.g. 2024-01-15).")


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip())
    except InvalidOperation:
        raise ValueError(f"Invalid decimal '{value}'.")


def _str(value: Any) -> str:
    """Coerce a cell value to a clean string."""
    if value is None:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# Template Generator
# ---------------------------------------------------------------------------

class ExcelTemplateGenerator:
    def generate_template(self, asset_type_id: int) -> openpyxl.Workbook:
        asset_type = AssetType.objects.select_related("category").get(pk=asset_type_id)
        spec_cols = [f"spec_{key}" for key in (asset_type.spec_schema or [])]
        all_cols = FIXED_COLUMNS + spec_cols

        wb = openpyxl.Workbook()
        self._build_data_sheet(wb, asset_type, all_cols, spec_cols)
        self._build_instructions_sheet(wb, asset_type, spec_cols)
        self._build_locations_sheet(wb)
        return wb

    # ------------------------------------------------------------------

    def _build_data_sheet(
        self,
        wb: openpyxl.Workbook,
        asset_type: AssetType,
        all_cols: list[str],
        spec_cols: list[str],
    ) -> None:
        ws = wb.active
        ws.title = "Data Entry"
        ws.freeze_panes = "A2"

        _DATE_COLS = {"purchase_date", "warranty_expiry", "amc_expiry"}

        # Header row — date columns get a format hint appended
        for col_idx, col_name in enumerate(all_cols, start=1):
            label = f"{col_name} (YYYY-MM-DD)" if col_name in _DATE_COLS else col_name
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN

        # Example row
        example = self._example_row(asset_type, spec_cols)
        for col_idx, col_name in enumerate(all_cols, start=1):
            cell = ws.cell(row=2, column=col_idx, value=example.get(col_name, ""))
            cell.font = _EXAMPLE_FONT
            if col_name in _DATE_COLS:
                cell.number_format = "@"  # text format prevents Excel date auto-conversion

        # Pre-format 300 data rows for date columns as text so Excel won't convert typed dates
        for col_idx, col_name in enumerate(all_cols, start=1):
            if col_name in _DATE_COLS:
                for row_i in range(3, 303):
                    ws.cell(row_i, col_idx).number_format = "@"

        # Auto-width: use max of header length and example value length
        for col_idx, col_name in enumerate(all_cols, start=1):
            label = f"{col_name} (YYYY-MM-DD)" if col_name in _DATE_COLS else col_name
            example_val = str(example.get(col_name, ""))
            width = max(len(label), len(example_val)) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 40)

        ws.row_dimensions[1].height = 28

    def _example_row(self, asset_type: AssetType, spec_cols: list[str]) -> dict[str, str]:
        # Pick the first active location path as an example if available
        first_loc = (
            Location.objects.filter(is_active=True)
            .select_related("parent__parent")
            .first()
        )
        loc_path = first_loc.full_path if first_loc else "Parliament Bhaban → 3rd Floor → NOC Room"

        row: dict[str, str] = {
            "asset_tag": "",
            "serial_number": "SN-EXAMPLE-001",
            "brand": "Dell",
            "model_name": "Latitude 5540",
            "purchase_date": "2024-01-15",
            "purchase_order": "PO/2024/001",
            "supplier": "Computer Source Ltd",
            "purchase_cost": "45000.00",
            "warranty_expiry": "2027-01-15",
            "amc_expiry": "2025-01-15",
            "storage_location_path": loc_path,
            "notes": "Example row — replace with actual data",
        }
        for spec_col in spec_cols:
            row[spec_col] = f"Example {spec_col[5:]}"  # strip "spec_" prefix
        return row

    def _build_instructions_sheet(
        self, wb: openpyxl.Workbook, asset_type: AssetType, spec_cols: list[str]
    ) -> None:
        ws = wb.create_sheet("Instructions")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 70

        row = 1
        ws.cell(row, 1, f"Template for: {asset_type.name}").font = _INSTR_HEADER_FONT
        ws.cell(row, 2, f"Category: {asset_type.category.name}").font = _INSTR_FONT
        row += 1

        ws.cell(row, 1, "Instructions").font = _INSTR_HEADER_FONT
        row += 1
        ws.cell(row, 1, "1. Fill in the 'Data Entry' sheet.").font = _INSTR_FONT
        row += 1
        ws.cell(row, 1, "2. Leave asset_tag blank — system auto-generates it.").font = _INSTR_FONT
        row += 1
        ws.cell(row, 1, "3. Copy location paths from the 'Valid Locations' sheet.").font = _INSTR_FONT
        row += 1
        ws.cell(row, 1, "4. Do not edit column headers.").font = _INSTR_FONT
        row += 2

        ws.cell(row, 1, "Column Reference").font = _INSTR_HEADER_FONT
        row += 1

        for col_name in FIXED_COLUMNS:
            desc = COLUMN_DESCRIPTIONS.get(col_name, "")
            ws.cell(row, 1, col_name).font = Font(bold=True, name="Calibri", size=10)
            ws.cell(row, 2, desc).font = _INSTR_FONT
            row += 1

        if spec_cols:
            ws.cell(row, 1, f"Spec columns ({len(spec_cols)})").font = _INSTR_HEADER_FONT
            row += 1
            for col_name in spec_cols:
                ws.cell(row, 1, col_name).font = Font(bold=True, name="Calibri", size=10)
                ws.cell(row, 2, "Free-text specification value. Optional.").font = _INSTR_FONT
                row += 1

    def _build_locations_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Valid Locations")
        ws.column_dimensions["A"].width = 70

        ws.cell(1, 1, "Valid Location Paths (copy-paste into storage_location_path column)").font = Font(
            bold=True, name="Calibri", size=11, color="006633"
        )

        locations = (
            Location.objects.filter(is_active=True)
            .select_related("parent__parent")
            .order_by("level_type", "name")
        )
        for row_idx, loc in enumerate(locations, start=2):
            ws.cell(row_idx, 1, loc.full_path).font = _INSTR_FONT


# ---------------------------------------------------------------------------
# Import Validator
# ---------------------------------------------------------------------------

class ExcelImportValidator:
    def validate(self, file_obj: Any, asset_type_id: int) -> list[dict]:
        asset_type = AssetType.objects.select_related("category").get(pk=asset_type_id)
        spec_keys = list(asset_type.spec_schema or [])
        expected_spec_cols = {f"spec_{k}" for k in spec_keys}

        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot read Excel file: {exc}") from exc

        # Prefer the "Data Entry" sheet; fall back to the first sheet
        ws = wb["Data Entry"] if "Data Entry" in wb.sheetnames else wb.active

        # Build column index map from header row
        # Strip format hints like " (YYYY-MM-DD)" added by the template generator
        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map: dict[str, int] = {}
        for idx, header in enumerate(header_row, start=1):
            if header:
                raw = str(header).strip()
                # Strip trailing " (FORMAT)" hint if present
                clean = raw.split(" (")[0].strip()
                col_map[clean] = idx

        # Pre-build location lookup once
        location_lookup = _build_location_lookup()

        # Existing asset tags for uniqueness check
        existing_tags: set[str] = set(
            AssetItem.objects.values_list("asset_tag", flat=True)
        )
        # Track tags seen in this batch to catch intra-batch duplicates
        batch_tags: set[str] = set()

        results: list[dict] = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {
                col_name: _str(ws.cell(row_idx, col_idx).value)
                for col_name, col_idx in col_map.items()
            }

            # Skip completely empty rows
            if not any(v for v in row_data.values()):
                continue

            errors: list[str] = []
            warnings: list[str] = []
            clean: dict[str, Any] = {}

            # --- Required fields ---
            brand = row_data.get("brand", "")
            if not brand:
                errors.append("brand is required.")
            else:
                clean["brand"] = brand

            model_name = row_data.get("model_name", "")
            if not model_name:
                errors.append("model_name is required.")
            else:
                clean["model_name"] = model_name

            # --- Serial number (warning if blank) ---
            serial = row_data.get("serial_number", "")
            clean["serial_number"] = serial
            if not serial:
                warnings.append("serial_number is blank — consider adding it for traceability.")

            # --- asset_tag uniqueness ---
            tag = row_data.get("asset_tag", "")
            clean["asset_tag"] = tag
            if tag:
                if tag in existing_tags or tag in batch_tags:
                    errors.append(f"asset_tag '{tag}' already exists.")
                else:
                    batch_tags.add(tag)

            # --- Date fields ---
            for date_field in ("purchase_date", "warranty_expiry", "amc_expiry"):
                raw = row_data.get(date_field, "")
                try:
                    parsed = _parse_date(raw if raw else None)
                    clean[date_field] = parsed.isoformat() if parsed else None
                except ValueError as exc:
                    errors.append(str(exc))
                    clean[date_field] = None

            # --- Decimal fields ---
            raw_cost = row_data.get("purchase_cost", "")
            try:
                parsed_cost = _parse_decimal(raw_cost if raw_cost else None)
                clean["purchase_cost"] = str(parsed_cost) if parsed_cost is not None else None
            except ValueError as exc:
                errors.append(str(exc))
                clean["purchase_cost"] = None

            # --- Storage location ---
            loc_path_raw = row_data.get("storage_location_path", "")
            clean["storage_location_id"] = None
            if loc_path_raw:
                key = _normalize_location_path(loc_path_raw)
                loc_id = location_lookup.get(key)
                if loc_id is None:
                    errors.append(
                        f"Location path '{loc_path_raw}' not found. "
                        "Copy-paste from the Valid Locations sheet."
                    )
                else:
                    clean["storage_location_id"] = loc_id

            # --- Simple string fields ---
            for field in ("purchase_order", "supplier", "notes"):
                clean[field] = row_data.get(field, "")

            # --- Spec columns ---
            specs: dict[str, str] = {}
            for spec_key in spec_keys:
                col_name = f"spec_{spec_key}"
                if col_name in row_data:
                    specs[spec_key] = row_data[col_name]
            clean["specifications"] = specs

            # Determine status
            if errors:
                status = "error"
            elif warnings:
                status = "warning"
            else:
                status = "valid"

            results.append(
                {
                    "row": row_idx,
                    "data": clean,
                    "errors": errors,
                    "warnings": warnings,
                    "status": status,
                }
            )

        return results


# ---------------------------------------------------------------------------
# Import Executor
# ---------------------------------------------------------------------------

class ExcelImportExecutor:
    def execute(
        self,
        validated_rows: list[dict],
        asset_type_id: int,
        user: Any,
    ) -> dict[str, Any]:
        asset_type = AssetType.objects.get(pk=asset_type_id)
        year = timezone.now().year

        importable = [r for r in validated_rows if r["status"] in ("valid", "warning")]
        skipped = len(validated_rows) - len(importable)

        tag_counters: dict[str, int] = {}
        created_count = 0
        errors: list[str] = []

        try:
            with transaction.atomic():
                for row in importable:
                    data = row["data"]
                    try:
                        asset_tag = data.get("asset_tag") or self._next_tag(
                            asset_type, year, tag_counters
                        )
                        AssetItem.objects.create(
                            asset_tag=asset_tag,
                            asset_type=asset_type,
                            serial_number=data.get("serial_number", ""),
                            brand=data["brand"],
                            model_name=data["model_name"],
                            specifications=data.get("specifications", {}),
                            status=AssetItem.Status.IN_STOCK,
                            storage_location_id=data.get("storage_location_id"),
                            purchase_date=data.get("purchase_date"),
                            purchase_order=data.get("purchase_order", ""),
                            supplier=data.get("supplier", ""),
                            purchase_cost=data.get("purchase_cost"),
                            warranty_expiry=data.get("warranty_expiry"),
                            amc_expiry=data.get("amc_expiry"),
                            notes=data.get("notes", ""),
                            created_by=user,
                        )
                        created_count += 1
                    except (IntegrityError, Exception) as exc:
                        errors.append(f"Row {row['row']}: {exc}")
                        raise  # trigger rollback of the whole batch
        except Exception:
            return {"created": 0, "skipped": skipped, "errors": errors}

        return {"created": created_count, "skipped": skipped, "errors": []}

    def _next_tag(
        self, asset_type: AssetType, year: int, counters: dict[str, int]
    ) -> str:
        prefix = _get_type_prefix(asset_type.name)
        key = f"{prefix}-{year}"
        if key not in counters:
            # Seed from the DB max so the batch picks up where we left off
            existing = AssetItem.objects.filter(
                asset_tag__startswith=f"{key}-"
            ).values_list("asset_tag", flat=True)
            max_seq = 0
            for tag in existing:
                try:
                    seq = int(tag[len(key) + 1:])
                    max_seq = max(max_seq, seq)
                except ValueError:
                    pass
            counters[key] = max_seq
        counters[key] += 1
        return f"{key}-{counters[key]:04d}"
