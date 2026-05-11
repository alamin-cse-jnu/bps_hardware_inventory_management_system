import io

import openpyxl
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.test import TestCase
from django.utils import timezone

from locations.models import Location

from .models import AssetCategory, AssetComponent, AssetItem, AssetType
from .services.excel_import import (
    FIXED_COLUMNS,
    ExcelImportExecutor,
    ExcelImportValidator,
    ExcelTemplateGenerator,
)

User = get_user_model()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_category(name="Computing") -> AssetCategory:
    return AssetCategory.objects.create(name=name)


def make_type(category=None, name="LAPTOP", has_components=False) -> AssetType:
    if category is None:
        category = make_category()
    return AssetType.objects.create(
        category=category,
        name=name,
        has_components=has_components,
        spec_schema=["cpu", "ram"] if has_components else ["cpu"],
    )


def make_item(asset_type=None, asset_tag="PC-2024-0001", status=AssetItem.Status.IN_STOCK) -> AssetItem:
    if asset_type is None:
        asset_type = make_type()
    return AssetItem.objects.create(
        asset_tag=asset_tag,
        asset_type=asset_type,
        brand="Dell",
        model_name="Latitude 5540",
        status=status,
    )


# ---------------------------------------------------------------------------
# State machine — valid transitions
# ---------------------------------------------------------------------------

class ValidTransitionTests(TestCase):
    """Every transition listed in CLAUDE.md must succeed."""

    def _item(self, status):
        return make_item(status=status, asset_tag=f"TAG-{status}")

    def test_in_stock_to_assigned(self):
        item = self._item(AssetItem.Status.IN_STOCK)
        item.change_status(AssetItem.Status.ASSIGNED)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.ASSIGNED)

    def test_in_stock_to_maintenance(self):
        item = self._item(AssetItem.Status.IN_STOCK)
        item.change_status(AssetItem.Status.MAINTENANCE)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.MAINTENANCE)

    def test_in_stock_to_disposed(self):
        item = self._item(AssetItem.Status.IN_STOCK)
        item.change_status(AssetItem.Status.DISPOSED)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.DISPOSED)

    def test_assigned_to_in_stock(self):
        item = self._item(AssetItem.Status.ASSIGNED)
        item.change_status(AssetItem.Status.IN_STOCK)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.IN_STOCK)

    def test_assigned_to_maintenance(self):
        item = self._item(AssetItem.Status.ASSIGNED)
        item.change_status(AssetItem.Status.MAINTENANCE)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.MAINTENANCE)

    def test_assigned_to_lost(self):
        item = self._item(AssetItem.Status.ASSIGNED)
        item.change_status(AssetItem.Status.LOST)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.LOST)

    def test_assigned_to_damaged(self):
        item = self._item(AssetItem.Status.ASSIGNED)
        item.change_status(AssetItem.Status.DAMAGED)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.DAMAGED)

    def test_assigned_to_disposed(self):
        item = self._item(AssetItem.Status.ASSIGNED)
        item.change_status(AssetItem.Status.DISPOSED)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.DISPOSED)

    def test_maintenance_to_in_stock(self):
        item = self._item(AssetItem.Status.MAINTENANCE)
        item.change_status(AssetItem.Status.IN_STOCK)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.IN_STOCK)

    def test_maintenance_to_disposed(self):
        item = self._item(AssetItem.Status.MAINTENANCE)
        item.change_status(AssetItem.Status.DISPOSED)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.DISPOSED)

    def test_lost_to_in_stock(self):
        item = self._item(AssetItem.Status.LOST)
        item.change_status(AssetItem.Status.IN_STOCK)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.IN_STOCK)

    def test_lost_to_disposed(self):
        item = self._item(AssetItem.Status.LOST)
        item.change_status(AssetItem.Status.DISPOSED)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.DISPOSED)

    def test_damaged_to_in_stock(self):
        item = self._item(AssetItem.Status.DAMAGED)
        item.change_status(AssetItem.Status.IN_STOCK)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.IN_STOCK)

    def test_damaged_to_maintenance(self):
        item = self._item(AssetItem.Status.DAMAGED)
        item.change_status(AssetItem.Status.MAINTENANCE)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.MAINTENANCE)

    def test_damaged_to_disposed(self):
        item = self._item(AssetItem.Status.DAMAGED)
        item.change_status(AssetItem.Status.DISPOSED)
        item.refresh_from_db()
        self.assertEqual(item.status, AssetItem.Status.DISPOSED)


# ---------------------------------------------------------------------------
# State machine — invalid transitions
# ---------------------------------------------------------------------------

class InvalidTransitionTests(TestCase):
    """Transitions not in CLAUDE.md must raise ValidationError."""

    def _assert_invalid(self, from_status, to_status):
        item = make_item(
            status=from_status,
            asset_tag=f"TAG-{from_status}-{to_status}",
        )
        with self.assertRaises(ValidationError):
            item.change_status(to_status)

    def test_in_stock_cannot_go_to_lost(self):
        self._assert_invalid(AssetItem.Status.IN_STOCK, AssetItem.Status.LOST)

    def test_in_stock_cannot_go_to_damaged(self):
        self._assert_invalid(AssetItem.Status.IN_STOCK, AssetItem.Status.DAMAGED)

    def test_maintenance_cannot_go_to_assigned(self):
        self._assert_invalid(AssetItem.Status.MAINTENANCE, AssetItem.Status.ASSIGNED)

    def test_maintenance_cannot_go_to_lost(self):
        self._assert_invalid(AssetItem.Status.MAINTENANCE, AssetItem.Status.LOST)

    def test_maintenance_cannot_go_to_damaged(self):
        self._assert_invalid(AssetItem.Status.MAINTENANCE, AssetItem.Status.DAMAGED)

    def test_lost_cannot_go_to_assigned(self):
        self._assert_invalid(AssetItem.Status.LOST, AssetItem.Status.ASSIGNED)

    def test_lost_cannot_go_to_maintenance(self):
        self._assert_invalid(AssetItem.Status.LOST, AssetItem.Status.MAINTENANCE)

    def test_lost_cannot_go_to_damaged(self):
        self._assert_invalid(AssetItem.Status.LOST, AssetItem.Status.DAMAGED)

    def test_disposed_cannot_go_to_in_stock(self):
        self._assert_invalid(AssetItem.Status.DISPOSED, AssetItem.Status.IN_STOCK)

    def test_disposed_cannot_go_to_assigned(self):
        self._assert_invalid(AssetItem.Status.DISPOSED, AssetItem.Status.ASSIGNED)

    def test_disposed_cannot_go_to_maintenance(self):
        self._assert_invalid(AssetItem.Status.DISPOSED, AssetItem.Status.MAINTENANCE)

    def test_disposed_cannot_go_to_lost(self):
        self._assert_invalid(AssetItem.Status.DISPOSED, AssetItem.Status.LOST)

    def test_disposed_cannot_go_to_damaged(self):
        self._assert_invalid(AssetItem.Status.DISPOSED, AssetItem.Status.DAMAGED)


# ---------------------------------------------------------------------------
# AssetComponent validation
# ---------------------------------------------------------------------------

class AssetComponentValidationTests(TestCase):
    def setUp(self):
        cat = make_category()
        self.pc_type = make_type(cat, "PC_SET", has_components=True)
        self.laptop_type = make_type(cat, "LAPTOP", has_components=False)
        self.pc = make_item(self.pc_type, "PC-001")
        self.laptop = make_item(self.laptop_type, "LT-001")

    def test_component_allowed_on_pc_set(self):
        comp = AssetComponent(
            parent_asset=self.pc,
            component_type=AssetComponent.ComponentType.MONITOR,
            brand="Samsung",
        )
        comp.full_clean()  # must not raise
        comp.save()
        self.assertEqual(comp.parent_asset, self.pc)

    def test_component_rejected_on_non_component_type(self):
        comp = AssetComponent(
            parent_asset=self.laptop,
            component_type=AssetComponent.ComponentType.RAM,
        )
        with self.assertRaises(ValidationError) as ctx:
            comp.full_clean()
        self.assertIn("parent_asset", ctx.exception.message_dict)

    def test_removed_component_kept_in_db(self):
        """Removed components stay in DB with is_active=False (architectural decision #1)."""
        comp = AssetComponent.objects.create(
            parent_asset=self.pc,
            component_type=AssetComponent.ComponentType.RAM,
            is_active=True,
        )
        comp.is_active = False
        comp.removed_at = timezone.now()
        comp.removal_reason = "Upgraded to 16GB"
        comp.save()

        self.assertIsNotNone(AssetComponent.objects.filter(pk=comp.pk).first())
        self.assertFalse(AssetComponent.objects.get(pk=comp.pk).is_active)


# ---------------------------------------------------------------------------
# Soft delete
# ---------------------------------------------------------------------------

class SoftDeleteTests(TestCase):
    def test_soft_delete_sets_flags(self):
        item = make_item(asset_tag="TAG-SOFT-DEL")
        self.assertFalse(item.is_deleted)
        self.assertIsNone(item.deleted_at)

        item.soft_delete()
        item.refresh_from_db()

        self.assertTrue(item.is_deleted)
        self.assertIsNotNone(item.deleted_at)

    def test_soft_delete_does_not_remove_row(self):
        item = make_item(asset_tag="TAG-SOFT-KEEP")
        pk = item.pk
        item.soft_delete()
        self.assertTrue(AssetItem.objects.filter(pk=pk).exists())


# ---------------------------------------------------------------------------
# asset_tag uniqueness
# ---------------------------------------------------------------------------

class AssetTagUniquenessTests(TestCase):
    def test_duplicate_asset_tag_raises_integrity_error(self):
        make_item(asset_tag="UNIQUE-001")
        from django.db import IntegrityError
        with self.assertRaises(IntegrityError):
            # Bypass full_clean to hit the DB constraint directly
            AssetItem.objects.create(
                asset_tag="UNIQUE-001",
                asset_type=make_type(make_category("Networking"), "SWITCH"),
                brand="Cisco",
                model_name="SG350",
            )


# ---------------------------------------------------------------------------
# is_assignable property
# ---------------------------------------------------------------------------

class IsAssignableTests(TestCase):
    def test_in_stock_not_deleted_is_assignable(self):
        item = make_item(status=AssetItem.Status.IN_STOCK)
        self.assertTrue(item.is_assignable)

    def test_assigned_is_not_assignable(self):
        item = make_item(status=AssetItem.Status.ASSIGNED, asset_tag="TAG-ASGN")
        self.assertFalse(item.is_assignable)

    def test_in_stock_but_deleted_is_not_assignable(self):
        item = make_item(status=AssetItem.Status.IN_STOCK, asset_tag="TAG-DEL")
        item.soft_delete()
        self.assertFalse(item.is_assignable)


# ===========================================================================
# Excel Import Tests
# ===========================================================================

# ---------------------------------------------------------------------------
# Shared fixtures for import tests
# ---------------------------------------------------------------------------

def make_import_category() -> AssetCategory:
    return AssetCategory.objects.create(name="ImportCat")


def make_import_type(spec_schema=None) -> AssetType:
    if spec_schema is None:
        spec_schema = ["cpu", "ram"]
    cat = make_import_category()
    return AssetType.objects.create(
        category=cat,
        name="LAPTOP",
        spec_schema=spec_schema,
    )


def make_location_hierarchy():
    building = Location.objects.create(name="Parliament Bhaban", level_type=Location.LevelType.BUILDING)
    floor = Location.objects.create(name="3rd Floor", level_type=Location.LevelType.FLOOR, parent=building)
    room = Location.objects.create(name="NOC Room", level_type=Location.LevelType.ROOM, parent=floor)
    return building, floor, room


def make_excel_file(headers: list, rows: list) -> io.BytesIO:
    """Build a minimal .xlsx with a 'Data Entry' sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Entry"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _base_headers(asset_type: AssetType) -> list:
    spec_cols = [f"spec_{k}" for k in (asset_type.spec_schema or [])]
    return FIXED_COLUMNS + spec_cols


def _valid_row_values(location_path: str = "") -> list:
    """Values matching the FIXED_COLUMNS order + 2 spec cols (cpu, ram)."""
    return [
        "",               # asset_tag (blank → auto-generate)
        "SN-TEST-001",    # serial_number
        "Dell",           # brand
        "Latitude 5540",  # model_name
        "2024-01-15",     # purchase_date
        "PO-001",         # purchase_order
        "Tech Supply",    # supplier
        "45000.00",       # purchase_cost
        "2027-01-15",     # warranty_expiry
        "2025-01-15",     # amc_expiry
        location_path,    # storage_location_path
        "Test notes",     # notes
        "Intel i7",       # spec_cpu
        "16GB",           # spec_ram
    ]


def make_import_user():
    return User.objects.create_user(username="importer", password="pass")


# ---------------------------------------------------------------------------
# ExcelTemplateGenerator tests
# ---------------------------------------------------------------------------

class ExcelTemplateGeneratorTests(TestCase):
    def setUp(self):
        self.asset_type = make_import_type(spec_schema=["cpu", "ram", "storage"])

    def test_template_has_three_sheets(self):
        wb = ExcelTemplateGenerator().generate_template(self.asset_type.pk)
        self.assertEqual(wb.sheetnames, ["Data Entry", "Instructions", "Valid Locations"])

    def test_template_fixed_columns_in_data_entry(self):
        wb = ExcelTemplateGenerator().generate_template(self.asset_type.pk)
        ws = wb["Data Entry"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
        for col in FIXED_COLUMNS:
            self.assertIn(col, headers, f"Missing fixed column: {col}")

    def test_template_dynamic_spec_columns(self):
        wb = ExcelTemplateGenerator().generate_template(self.asset_type.pk)
        ws = wb["Data Entry"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
        for spec_key in self.asset_type.spec_schema:
            self.assertIn(f"spec_{spec_key}", headers, f"Missing spec column: spec_{spec_key}")

    def test_template_has_example_row(self):
        wb = ExcelTemplateGenerator().generate_template(self.asset_type.pk)
        ws = wb["Data Entry"]
        # Row 2 should have at least brand filled in
        row2_vals = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in row2_vals if v]
        self.assertGreater(len(non_empty), 0, "Example row is completely empty")

    def test_template_valid_locations_sheet_lists_active_locations(self):
        building, floor, room = make_location_hierarchy()
        wb = ExcelTemplateGenerator().generate_template(self.asset_type.pk)
        ws = wb["Valid Locations"]
        paths = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value]
        self.assertIn(building.full_path, paths)
        self.assertIn(floor.full_path, paths)
        self.assertIn(room.full_path, paths)


# ---------------------------------------------------------------------------
# ExcelImportValidator tests
# ---------------------------------------------------------------------------

class ExcelImportValidatorTests(TestCase):
    def setUp(self):
        self.asset_type = make_import_type(spec_schema=["cpu", "ram"])
        self.building, self.floor, self.room = make_location_hierarchy()
        self.headers = _base_headers(self.asset_type)

    def _validate(self, rows):
        f = make_excel_file(self.headers, rows)
        return ExcelImportValidator().validate(f, self.asset_type.pk)

    def test_valid_row_passes(self):
        results = self._validate([_valid_row_values(self.room.full_path)])
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["status"], "valid")
        self.assertEqual(results[0]["errors"], [])

    def test_missing_brand_fails(self):
        row = _valid_row_values()
        row[2] = ""  # brand is index 2
        results = self._validate([row])
        self.assertEqual(results[0]["status"], "error")
        self.assertTrue(any("brand" in e for e in results[0]["errors"]))

    def test_missing_model_fails(self):
        row = _valid_row_values()
        row[3] = ""  # model_name is index 3
        results = self._validate([row])
        self.assertEqual(results[0]["status"], "error")
        self.assertTrue(any("model_name" in e for e in results[0]["errors"]))

    def test_blank_serial_produces_warning(self):
        row = _valid_row_values()
        row[1] = ""  # serial_number is index 1
        results = self._validate([row])
        self.assertIn(results[0]["status"], ("warning",))
        self.assertTrue(any("serial_number" in w for w in results[0]["warnings"]))

    def test_duplicate_asset_tag_fails(self):
        # Pre-create an asset with tag "TAG-DUP-001"
        make_item(asset_type=self.asset_type, asset_tag="TAG-DUP-001")
        row = _valid_row_values()
        row[0] = "TAG-DUP-001"  # asset_tag
        results = self._validate([row])
        self.assertEqual(results[0]["status"], "error")
        self.assertTrue(any("TAG-DUP-001" in e for e in results[0]["errors"]))

    def test_invalid_location_path_fails(self):
        row = _valid_row_values("Completely → Wrong → Path")
        results = self._validate([row])
        self.assertEqual(results[0]["status"], "error")
        self.assertTrue(any("not found" in e.lower() for e in results[0]["errors"]))

    def test_valid_location_path_resolves(self):
        row = _valid_row_values(self.room.full_path)
        results = self._validate([row])
        self.assertEqual(results[0]["data"]["storage_location_id"], self.room.pk)

    def test_empty_rows_are_skipped(self):
        # Mix: one valid row, two empty rows
        valid = _valid_row_values()
        results = self._validate([valid, ["", "", "", "", "", "", "", "", "", "", "", "", "", ""], [None] * 14])
        self.assertEqual(len(results), 1)

    def test_invalid_date_format_fails(self):
        row = _valid_row_values()
        row[4] = "15/01/2024"  # purchase_date — wrong format
        results = self._validate([row])
        self.assertEqual(results[0]["status"], "error")
        self.assertTrue(any("date" in e.lower() for e in results[0]["errors"]))

    def test_invalid_cost_fails(self):
        row = _valid_row_values()
        row[7] = "not-a-number"  # purchase_cost
        results = self._validate([row])
        self.assertEqual(results[0]["status"], "error")
        self.assertTrue(any("decimal" in e.lower() for e in results[0]["errors"]))


# ---------------------------------------------------------------------------
# ExcelImportExecutor tests
# ---------------------------------------------------------------------------

class ExcelImportExecutorTests(TestCase):
    def setUp(self):
        self.asset_type = make_import_type(spec_schema=["cpu", "ram"])
        self.user = make_import_user()

    def _valid_row_dict(self, asset_tag="", serial="SN-001", brand="Dell", model="D1", row_num=2):
        return {
            "row": row_num,
            "status": "valid",
            "errors": [],
            "warnings": [],
            "data": {
                "asset_tag": asset_tag,
                "serial_number": serial,
                "brand": brand,
                "model_name": model,
                "purchase_date": None,
                "purchase_order": "",
                "supplier": "",
                "purchase_cost": None,
                "warranty_expiry": None,
                "amc_expiry": None,
                "storage_location_id": None,
                "notes": "",
                "specifications": {"cpu": "i7", "ram": "16GB"},
            },
        }

    def test_creates_assets_from_valid_rows(self):
        rows = [self._valid_row_dict(asset_tag="LAP-2024-TEST1")]
        result = ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["errors"], [])
        self.assertTrue(AssetItem.objects.filter(asset_tag="LAP-2024-TEST1").exists())

    def test_auto_generates_asset_tag_format(self):
        rows = [self._valid_row_dict()]  # asset_tag blank
        result = ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        self.assertEqual(result["created"], 1)
        item = AssetItem.objects.first()
        year = timezone.now().year
        self.assertRegex(item.asset_tag, rf"^LAP-{year}-\d{{4}}$")

    def test_asset_tag_sequence_increments_within_batch(self):
        rows = [
            self._valid_row_dict(row_num=2, serial="SN-A"),
            self._valid_row_dict(row_num=3, serial="SN-B"),
        ]
        result = ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        self.assertEqual(result["created"], 2)
        year = timezone.now().year
        tags = sorted(AssetItem.objects.values_list("asset_tag", flat=True))
        self.assertEqual(tags[0], f"LAP-{year}-0001")
        self.assertEqual(tags[1], f"LAP-{year}-0002")

    def test_asset_tag_sequence_continues_from_db(self):
        # Pre-existing asset with sequence 0005
        year = timezone.now().year
        make_item(asset_type=self.asset_type, asset_tag=f"LAP-{year}-0005")
        rows = [self._valid_row_dict()]
        ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        self.assertTrue(AssetItem.objects.filter(asset_tag=f"LAP-{year}-0006").exists())

    def test_sets_created_by_user(self):
        rows = [self._valid_row_dict(asset_tag="LAP-USR-001")]
        ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        item = AssetItem.objects.get(asset_tag="LAP-USR-001")
        self.assertEqual(item.created_by, self.user)

    def test_status_set_to_in_stock(self):
        rows = [self._valid_row_dict(asset_tag="LAP-STOCK-001")]
        ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        item = AssetItem.objects.get(asset_tag="LAP-STOCK-001")
        self.assertEqual(item.status, AssetItem.Status.IN_STOCK)

    def test_skips_error_rows(self):
        rows = [
            self._valid_row_dict(asset_tag="LAP-OK-001"),
            {**self._valid_row_dict(asset_tag="LAP-BAD-001", row_num=3), "status": "error"},
        ]
        result = ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["skipped"], 1)
        self.assertFalse(AssetItem.objects.filter(asset_tag="LAP-BAD-001").exists())

    def test_transaction_rollback_on_duplicate_tag_within_batch(self):
        """Two rows with the same explicit tag: first succeeds, second fails → rollback both."""
        rows = [
            self._valid_row_dict(asset_tag="LAP-DUP-001", row_num=2),
            self._valid_row_dict(asset_tag="LAP-DUP-001", row_num=3),  # duplicate
        ]
        result = ExcelImportExecutor().execute(rows, self.asset_type.pk, self.user)
        self.assertEqual(result["created"], 0)
        self.assertFalse(AssetItem.objects.filter(asset_tag="LAP-DUP-001").exists())
        self.assertGreater(len(result["errors"]), 0)

    def test_warning_rows_are_imported(self):
        row = {**self._valid_row_dict(asset_tag="LAP-WARN-001"), "status": "warning"}
        result = ExcelImportExecutor().execute([row], self.asset_type.pk, self.user)
        self.assertEqual(result["created"], 1)
        self.assertTrue(AssetItem.objects.filter(asset_tag="LAP-WARN-001").exists())
